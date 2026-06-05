"""
services/report_export_service.py
=================================
Batch export for teacher summary reports with concurrent processing.
"""

from __future__ import annotations

import asyncio
import logging
import os
from datetime import timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select, update, delete
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.core.database import AsyncSessionLocal
from app.core.time_utils import now_biz, today_biz, now_biz_dt_for_db, days_ago_biz
from app.models.models import SummaryReportExportJob, User
from app.services import analysis_service

logger = logging.getLogger(__name__)

EXPORT_DIR = Path(settings.export_dir)


def _now_local_str() -> str:
    return now_biz().strftime("%Y-%m-%d %H:%M:%S")


def _derive_class_code(student_no: str | None) -> str | None:
    return analysis_service.derive_class_code(student_no)


async def _get_students_by_class(db: AsyncSession, class_code: str | None) -> list[User]:
    stmt = select(User).where(User.role == "student")
    res = await db.execute(stmt)
    users = list(res.scalars().all())
    if not class_code:
        return users
    filtered = []
    for u in users:
        code = _derive_class_code(u.student_no)
        if code == class_code:
            filtered.append(u)
    return filtered


def _build_profile_row(profile: dict) -> list:
    def _score(key: str) -> int | None:
        v = profile.get(key) if isinstance(profile, dict) else None
        if isinstance(v, dict):
            return v.get("score")
        return None

    return [
        _score("code_understanding"),
        _score("new_tech_learning"),
        _score("communication"),
        _score("self_learning_and_frequency"),
        _score("tech_ethics_values"),
        _score("asks_direct_answers"),
    ]


def _build_profile_notes(profile: dict) -> str:
    def _notes(key: str) -> str:
        v = profile.get(key) if isinstance(profile, dict) else None
        if isinstance(v, dict):
            return v.get("notes", "") or ""
        return ""

    notes = [
        _notes("code_understanding"),
        _notes("new_tech_learning"),
        _notes("communication"),
        _notes("self_learning_and_frequency"),
        _notes("tech_ethics_values"),
        _notes("asks_direct_answers"),
    ]
    return " | ".join([n for n in notes if n])


def _write_excel(
    rows: list[list],
    class_code: str | None,
    course_name: str | None,
    teacher_name: str | None,
    school_name: str | None,
    include_text: bool,
    daily_rows: list[dict] | None = None,
) -> str:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "成绩导出"

    columns = [
        "序号",
        "学生姓名",
        "学号/工号",
        "学校",
        "班级",
        "代码理解(25%)",
        "学习新技术(25%)",
        "沟通能力(10%)",
        "自学能力/问次数(10%)",
        "技术伦理/价值观(20%)",
        "是否直接要答案(10%)",
        "综合成绩",
    ]
    if include_text:
        columns.append("评价")

    # 配置冻结窗格：冻结前两行及首列
    ws.freeze_panes = "B3"

    header_text = f"学校：{school_name or '-'}    课程：{course_name or '-'}    班级：{class_code or '-'}    任课教师：{teacher_name or '-'}  导出时间：{_now_local_str()}"
    ws.append([header_text])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    ws["A1"].fill = PatternFill("solid", fgColor="78A9F2")
    ws["A1"].font = Font(bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.append(columns)
    for col in range(1, len(columns) + 1):
        cell = ws.cell(row=2, column=col)
        cell.fill = PatternFill("solid", fgColor="CBDFFE")
        cell.font = Font(bold=True, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    data_fills = [
        PatternFill("solid", fgColor="F1F6FD"),
        PatternFill("solid", fgColor="FFFFFF"),
    ]

    # Sort rows by index (first column) to keep original sequence
    rows.sort(key=lambda x: x[0])

    for idx, row in enumerate(rows, start=1):
        ws.append(row)
        data_row_idx = idx + 2
        fill = data_fills[idx % 2]
        for col in range(1, len(columns) + 1):
            ws.cell(row=data_row_idx, column=col).fill = fill

    # 设置行高与列宽
    for r in range(1, len(rows) + 3):
        ws.row_dimensions[r].height = 30

    for col in range(1, len(columns) + 1):
        # Limited to 26 columns based on chr(64+col)
        ws.column_dimensions[chr(64 + col)].width = 18

    # 第二个工作表：每日分析（竖排为学生，横排为日期）
    _write_daily_analysis_sheet(wb, daily_rows or [])

    filename = f"summary_report_export_{now_biz().strftime('%Y%m%d_%H%M%S')}.xlsx"
    path = EXPORT_DIR / filename
    wb.save(path)
    return str(path)


def _write_daily_analysis_sheet(wb: Workbook, daily_rows: list[dict]) -> None:
    """
    写入「每日分析」工作表：每个学生一行，每个日期一列，单元格为当日分析文本。

    daily_rows 每项结构：
        {"idx": int, "name": str, "student_no": str, "daily": {date_str: analysis_text}}
    """
    ws = wb.create_sheet(title="每日分析")

    # 汇总所有出现过的日期（升序），作为日期列
    all_dates = sorted({d for entry in daily_rows for d in (entry.get("daily") or {}).keys()})

    base_columns = ["序号", "学生姓名", "学号/工号"]
    columns = base_columns + all_dates
    total_cols = len(columns)

    # 冻结前两行 + 前三列（学生信息列）
    ws.freeze_panes = "D2"

    # 表头
    ws.append(columns)
    for col in range(1, total_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = PatternFill("solid", fgColor="CBDFFE")
        cell.font = Font(bold=True, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_fills = [
        PatternFill("solid", fgColor="F1F6FD"),
        PatternFill("solid", fgColor="FFFFFF"),
    ]

    # 按序号排序，保持与成绩表一致的学生顺序
    sorted_rows = sorted(daily_rows, key=lambda x: x.get("idx", 0))

    for i, entry in enumerate(sorted_rows, start=1):
        daily = entry.get("daily") or {}
        row = [
            entry.get("idx", i),
            entry.get("name", ""),
            entry.get("student_no", ""),
        ]
        for d in all_dates:
            row.append(daily.get(d, ""))
        ws.append(row)

        data_row_idx = i + 1
        fill = data_fills[i % 2]
        ws.row_dimensions[data_row_idx].height = 120
        for col in range(1, total_cols + 1):
            cell = ws.cell(row=data_row_idx, column=col)
            cell.fill = fill
            # 学生信息列居中，分析内容列左上对齐并自动换行
            if col <= len(base_columns):
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # 列宽：信息列较窄，日期分析列较宽
    ws.column_dimensions[get_column_letter(1)].width = 8   # 序号
    ws.column_dimensions[get_column_letter(2)].width = 14  # 姓名
    ws.column_dimensions[get_column_letter(3)].width = 18  # 学号
    for col in range(len(base_columns) + 1, total_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 45

    ws.row_dimensions[1].height = 24


async def run_export_job(job_id: str) -> None:
    """
    运行导出任务，支持学生级别并发。
    """
    async with AsyncSessionLocal() as db:
        job = await db.get(SummaryReportExportJob, job_id)
        if not job or job.status != "pending":
            return

        job.status = "running"
        job.updated_at = now_biz_dt_for_db()
        await db.commit()

    # Re-fetch for processing
    async with AsyncSessionLocal() as db:
        job = await db.get(SummaryReportExportJob, job_id)
        try:
            users = await _get_students_by_class(db, job.class_code)
            job.total_count = len(users)
            await db.commit()

            if not users:
                job.status = "completed"
                await db.commit()
                return

            end_date = today_biz()
            start_date = end_date - timedelta(days=settings.max_report_days - 1)
            start_str = start_date.isoformat()
            end_str = end_date.isoformat()

            results: list[list] = []
            daily_results: list[dict] = []
            semaphore = asyncio.Semaphore(settings.export_concurrency)

            async def _process_student(user: User, idx: int):
                async with semaphore:
                    try:
                        # Use a dedicated session for each student to avoid cross-sharing issues
                        async with AsyncSessionLocal() as sub_db:
                            # 收集每日分析内容（用于「每日分析」工作表）
                            daily_analyses = await analysis_service.get_daily_analyses(
                                db=sub_db,
                                user_id=user.user_id,
                                start_date=start_str,
                                end_date=end_str,
                            )
                            if daily_analyses:
                                daily_results.append({
                                    "idx": idx,
                                    "name": user.real_name or "",
                                    "student_no": user.student_no or user.user_id,
                                    "daily": {r.date: (r.analysis_text or "") for r in daily_analyses},
                                })

                            report = await analysis_service.get_or_generate_summary_report(
                                db=sub_db,
                                user_id=user.user_id,
                                start_date=start_str,
                                end_date=end_str,
                                force=False,
                            )
                            report_json = report.report_json
                            total_score = report.total_score

                            profile = (report_json or {}).get("profile", {}) if isinstance(report_json, dict) else {}
                            profile_row = _build_profile_row(profile)
                            notes_text = _build_profile_notes(profile) if job.include_text_evaluation else None

                            class_code = _derive_class_code(user.student_no) or ""
                            row = [
                                idx,
                                user.real_name or "",
                                user.student_no or user.user_id,
                                job.school_name or "河北农业大学",
                                class_code,
                                *profile_row,
                                total_score,
                            ]
                            if notes_text is not None:
                                row.append(notes_text)
                            
                            results.append(row)
                            
                            # Atomic update progress
                            async with AsyncSessionLocal() as progress_db:
                                await progress_db.execute(
                                    update(SummaryReportExportJob)
                                    .where(SummaryReportExportJob.id == job_id)
                                    .values(
                                        completed_count=SummaryReportExportJob.completed_count + 1,
                                        updated_at=now_biz_dt_for_db()
                                    )
                                )
                                await progress_db.commit()
                    except ValueError as ex:
                        if str(ex) == "no_daily_analysis":
                            # 该学生在统计周期内无分析数据，属于正常缺省，静默跳过
                            logger.debug(f"[导出任务] 学生 {user.user_id} 暂无分析数据，已跳过")
                            async with AsyncSessionLocal() as progress_db:
                                await progress_db.execute(
                                    update(SummaryReportExportJob)
                                    .where(SummaryReportExportJob.id == job_id)
                                    .values(
                                        completed_count=SummaryReportExportJob.completed_count + 1,
                                        updated_at=now_biz_dt_for_db()
                                    )
                                )
                                await progress_db.commit()
                        else:
                            logger.error(f"[导出任务] 处理学生 {user.user_id} 失败: {ex}")
                            async with AsyncSessionLocal() as progress_db:
                                await progress_db.execute(
                                    update(SummaryReportExportJob)
                                    .where(SummaryReportExportJob.id == job_id)
                                    .values(
                                        failed_count=SummaryReportExportJob.failed_count + 1,
                                        updated_at=now_biz_dt_for_db()
                                    )
                                )
                                await progress_db.commit()
                    except Exception as ex:
                        logger.error(f"[导出任务] 处理学生 {user.user_id} 失败: {ex}")
                        async with AsyncSessionLocal() as progress_db:
                            await progress_db.execute(
                                update(SummaryReportExportJob)
                                .where(SummaryReportExportJob.id == job_id)
                                .values(
                                    failed_count=SummaryReportExportJob.failed_count + 1,
                                    updated_at=now_biz_dt_for_db()
                                )
                            )
                            await progress_db.commit()

            # Concurrent execution
            tasks = [_process_student(u, i) for i, u in enumerate(users, start=1)]
            await asyncio.gather(*tasks)

            # Finalize
            job = await db.get(SummaryReportExportJob, job_id)
            result_path = _write_excel(
                rows=results,
                class_code=job.class_code,
                course_name=job.course_name,
                teacher_name=job.teacher_name,
                school_name=job.school_name,
                include_text=job.include_text_evaluation,
                daily_rows=daily_results,
            )
            job.result_path = result_path
            job.status = "completed" if job.failed_count == 0 else "completed_with_errors"
            job.updated_at = now_biz_dt_for_db()
            await db.commit()

        except Exception as e:
            logger.error(f"[导出任务] 严重错误 job_id={job_id}: {e}", exc_info=True)
            job = await db.get(SummaryReportExportJob, job_id)
            job.status = "failed"
            job.error_message = str(e)
            job.updated_at = now_biz_dt_for_db()
            await db.commit()


async def reset_stale_jobs() -> int:
    """
    自愈：将所有处于 'running' 状态的任务重置为 'failed'。
    应用场景：程序异常崩溃重启后，修复卡死的状态。
    """
    async with AsyncSessionLocal() as db:
        stmt = (
            update(SummaryReportExportJob)
            .where(SummaryReportExportJob.status == "running")
            .values(
                status="failed",
                error_message="Server restarted during processing",
                updated_at=now_biz_dt_for_db()
            )
        )
        res = await db.execute(stmt)
        await db.commit()
        return res.rowcount


async def process_all_pending_jobs() -> None:
    """
    处理所有处于 'pending' 状态的任务。
    """
    async with AsyncSessionLocal() as db:
        stmt = select(SummaryReportExportJob.id).where(SummaryReportExportJob.status == "pending")
        res = await db.execute(stmt)
        job_ids = [str(jid) for jid in res.scalars().all()]
    
    if not job_ids:
        return

    logger.info(f"[任务队列] 发现 {len(job_ids)} 个待处理导出任务")
    for jid in job_ids:
        # Sequential processing of jobs, but internal students are concurrent
        await run_export_job(jid)


def cleanup_old_exports(days: int = 7) -> int:
    if not EXPORT_DIR.exists():
        return 0
    count = 0
    cutoff = days_ago_biz(days)
    for f in EXPORT_DIR.glob("*.xlsx"):
        if f.is_file():
            from datetime import datetime
            from app.core.config import CHINA_TZ
            mtime = datetime.fromtimestamp(f.stat().st_mtime, tz=CHINA_TZ)
            if mtime < cutoff:
                try:
                    f.unlink()
                    count += 1
                except Exception:
                    pass
    return count


async def cleanup_old_export_jobs(db: AsyncSession, days: int = 7) -> int:
    cutoff = days_ago_biz(days)
    stmt = delete(SummaryReportExportJob).where(
        SummaryReportExportJob.created_at < cutoff
    )
    res = await db.execute(stmt)
    await db.flush()
    return res.rowcount
